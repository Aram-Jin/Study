from bs4.dammit import encoding_res
import pandas as pd
import requests
from lxml import html
from tqdm import tqdm
from bs4 import BeautifulSoup
import csv
import openpyxl 
from openpyxl import Workbook, workbook

codes = '091970', '091970', '096640'    
price = []
for code in codes:  
    url = f'https://comp.fnguide.com/SVO2/ASP/SVD_FinanceRatio.asp?pGB=1&gicode=A{code}&cID=&MenuYn=Y&ReportGB=B&NewMenuID=104&stkGb=701'
    res = requests.get(url)
    data = res.text  
    soup = BeautifulSoup(data, "html.parser")
    for i in range(1,24):  
        aaa = soup.select_one("#p_grid1_"+ str(i))
        for j in aaa:
            if j.name == 'td.r':
                aaa.append
            else:
               pass 
        price.append(aaa)
        
        print(price) 
            #if j.name == 'td.r':
            #    data.append
            #else:
            #    pass 
            #price.append(data)
            
  
    
#1) 엑셀만들기    
# wb = openpyxl.Workbook()    

# #2) 엑셀 워크시트 만들기
# ws = wb.create_sheet('재무비율')

# #3) 데이터 추가하기
# ws['A1'] = '종목'
   
    
    
    
    
    
    
    
    
    
    
    
    
    
